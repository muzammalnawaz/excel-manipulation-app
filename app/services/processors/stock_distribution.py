from typing import Dict, List, Tuple, Optional
import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .base_processor import BaseProcessor
from ...core.exceptions import FileProcessingError

class StockDistributionProcessor(BaseProcessor):
    
    def validate_inputs(self, warehouse_path: str, sales_path: str) -> bool:
        if not warehouse_path or not sales_path:
            raise FileProcessingError("Both warehouse and sales file paths are required")
        return True
    
    def rename_columns(self, df: pd.DataFrame, column_mapping: Dict[int, str]) -> pd.DataFrame:
        columns = list(df.columns)
        new_columns = {}
        
        for position, new_name in column_mapping.items():
            if position < len(columns):
                new_columns[columns[position]] = new_name
        
        for i, col in enumerate(columns):
            if i not in column_mapping:
                new_columns[col] = col
        
        return df.rename(columns=new_columns)
    
    def get_warehouse_column_mapping(self) -> Dict[int, str]:
        return {
            0: 'Company Code',
            1: 'Company Code Name',
            2: 'Hierarchy Level 2 Code',
            3: 'Department Name',
            4: 'Hierarchy Level 1 Code', 
            5: 'Hierarchy Level 1 Description',
            6: 'Nesto MC Code',
            7: 'Nesto MC Description',
            8: 'Brand Code',
            9: 'Brand Description',
            10: 'Article Code',
            11: 'Article Description'
        }
    
    def get_sales_column_mapping(self) -> Dict[int, str]:
        return {
            0: 'Company Code',
            1: 'Company Code Name',
            2: 'Hierarchy Level 2 Code',
            3: 'Department Name',
            4: 'Hierarchy Level 1 Code',
            5: 'Hierarchy Level 1 Description', 
            6: 'Nesto MC Code',
            7: 'Nesto MC Description',
            8: 'Brand Code',
            9: 'Brand Description',
            10: 'Article Code',
            11: 'Article Description'
        }
    
    def find_quantity_column(self, df: pd.DataFrame, column_type: str = "warehouse") -> str:
        for col in df.columns:
            col_str = str(col).lower()
            if column_type == "warehouse":
                if 'total' in col_str and 'quantity' in col_str:
                    return col
            elif column_type == "sales_stock":
                if 'stock' in col_str and 'qty' in col_str and 'current' in col_str:
                    return col
            elif column_type == "sales_qty":
                if 'sales' in col_str and 'qty' in col_str:
                    return col
        return None
    
    def get_warehouse_quantity_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        quantity_columns = {}
        patterns = {
            'Quantity 0-30': ['quantity 0-30', 'qty 0-30', '0-30 days', '0-30'],
            'Quantity 31-60': ['quantity 31-60', 'qty 31-60', '31-60 days', '31-60'],
            'Quantity 61-90': ['quantity 61-90', 'qty 61-90', '61-90 days', '61-90'],
            'Quantity 91-180': ['quantity 91-180', 'qty 91-180', '91-180 days', '91-180'],
            'Quantity 181-365': ['quantity 181-365', 'qty 181-365', '181-365 days', '181-365'],
            'Quantity >365': ['quantity >365', 'qty >365', '>365 days', '>365'],
            'Quantity >180': ['quantity >180', 'qty >180', '>180 days', '>180']
        }
        
        for standard_name, search_patterns in patterns.items():
            for col in df.columns:
                col_str = str(col).lower().strip()
                for pattern in search_patterns:
                    if pattern in col_str:
                        quantity_columns[standard_name] = col
                        break
                if standard_name in quantity_columns:
                    break
        
        return quantity_columns
    
# Update stock_distribution.py process method with PROMINENT debugging:

    def process(self, warehouse_path: str, sales_path: str, 
                start_date: Optional[str] = None, end_date: Optional[str] = None) -> str:
        try:
            print("\n" + "🟩"*60)
            print("🔥 STOCK DISTRIBUTION PROCESSOR - MAIN PROCESS METHOD")
            print("🟩"*60)
            print(f"📁 warehouse_path: {warehouse_path}")
            print(f"📁 sales_path: {sales_path}")
            print(f"📅 RECEIVED start_date: '{start_date}' (type: {type(start_date)})")
            print(f"📅 RECEIVED end_date: '{end_date}' (type: {type(end_date)})")
            print("🟩"*60 + "\n")
            
            self.validate_inputs(warehouse_path, sales_path)
            
            print("Reading warehouse data from SAP BI file...")
            warehouse_df = self.read_excel_file(warehouse_path, "warehouse")
            
            print("Reading sales data from SAP BI file...")
            sales_df = self.read_excel_file(sales_path, "sales")
            
            if warehouse_df.empty:
                raise FileProcessingError("No valid warehouse data found")
            
            if sales_df.empty:
                raise FileProcessingError("No valid sales data found")
            
            print(f"Data loaded - Warehouse: {warehouse_df.shape}, Sales: {sales_df.shape}")
            
            warehouse_processed = self._process_warehouse_data(warehouse_df)
            sales_processed = self._process_sales_data(sales_df)
            
            print(f"Processed data - Warehouse: {warehouse_processed.shape}, Sales: {sales_processed.shape}")
            
            # THIS IS WHERE THE DATES GO TO MC PERFORMANCE CALCULATION
            print(f"\n🚀 CALLING _calculate_mc_performance with:")
            print(f"   start_date: '{start_date}'")
            print(f"   end_date: '{end_date}'")
            
            mc_performance = self._calculate_mc_performance(sales_processed, start_date, end_date)
            
            # Rest of processing...
            existing_distribution = self._calculate_existing_distributions(warehouse_processed, sales_processed, mc_performance)
            warehouse_remaining = self._update_warehouse_stock(warehouse_processed, existing_distribution)
            mc_stock_status = self._calculate_mc_stock_after_existing(sales_processed, existing_distribution, mc_performance)
            entirely_new_distribution = self._calculate_entirely_new_distributions(
                warehouse_remaining, mc_stock_status, mc_performance, sales_processed
            )
            warehouse_remaining = self._update_warehouse_stock(warehouse_remaining, entirely_new_distribution)
            mc_stock_status = self._calculate_mc_stock_after_distribution(mc_stock_status, entirely_new_distribution, mc_performance)
            site_new_distribution = self._calculate_site_new_distributions(
                warehouse_remaining, sales_processed, mc_stock_status, mc_performance
            )
            all_distributions = existing_distribution + entirely_new_distribution + site_new_distribution
            
            output_data = self._create_output_data(
                all_distributions, existing_distribution, entirely_new_distribution, 
                site_new_distribution, warehouse_processed, sales_processed, start_date, end_date, mc_performance
            )
            
            output_path = self.get_output_path("stock_distribution")
            self._save_excel_with_colors(output_data, output_path)
            
            print(f"Stock distribution analysis completed successfully")
            print(f"Existing: {len(existing_distribution)}, Entirely New: {len(entirely_new_distribution)}, Site New: {len(site_new_distribution)}")
            
            return output_path
            
        except Exception as e:
            if isinstance(e, FileProcessingError):
                raise e
            raise FileProcessingError(f"Error processing stock distribution: {str(e)}")
    
    def _process_warehouse_data(self, warehouse_df: pd.DataFrame) -> pd.DataFrame:
        warehouse_df = self.clean_dataframe(warehouse_df)
        warehouse_renamed = self.rename_columns(warehouse_df, self.get_warehouse_column_mapping())
        
        warehouse_clean = warehouse_renamed.dropna(subset=['Article Description'])
        
        warehouse_processed = pd.DataFrame()
        warehouse_processed['company_code'] = warehouse_clean['Company Code']
        warehouse_processed['department'] = warehouse_clean['Department Name']
        warehouse_processed['hierarchy_level1'] = warehouse_clean['Hierarchy Level 1 Description']
        warehouse_processed['nesto_mc'] = warehouse_clean['Nesto MC Description']
        warehouse_processed['brand'] = warehouse_clean['Brand Description']
        warehouse_processed['article_code'] = warehouse_clean['Article Code']
        warehouse_processed['article_description'] = warehouse_clean['Article Description']
        warehouse_processed['article_key'] = warehouse_clean['Article Code'].astype(str)
        
        qty_col = self.find_quantity_column(warehouse_clean, "warehouse")
        if qty_col:
            warehouse_processed['total_quantity'] = pd.to_numeric(warehouse_clean[qty_col], errors='coerce').fillna(0)
            warehouse_processed['remaining_quantity'] = warehouse_processed['total_quantity'].copy()
            print(f"Found warehouse quantity column: {qty_col}")
        else:
            raise FileProcessingError("Could not find total quantity column in warehouse data")
        
        quantity_columns = self.get_warehouse_quantity_columns(warehouse_clean)
        for standard_name, original_col in quantity_columns.items():
            warehouse_processed[standard_name] = pd.to_numeric(warehouse_clean[original_col], errors='coerce').fillna(0)
            print(f"Found warehouse aging column: {original_col} -> {standard_name}")
        
        warehouse_processed = warehouse_processed[
            (warehouse_processed['article_code'].notna()) & 
            (warehouse_processed['article_code'] != '') & 
            (warehouse_processed['total_quantity'] > 0)
        ]
        
        duplicate_articles = warehouse_processed['article_key'].duplicated().sum()
        if duplicate_articles > 0:
            print(f"Found {duplicate_articles} duplicate articles in warehouse data")
            agg_dict = {
                'company_code': 'first',
                'department': 'first',
                'hierarchy_level1': 'first', 
                'nesto_mc': 'first',
                'brand': 'first',
                'article_code': 'first',
                'article_description': 'first',
                'total_quantity': 'sum',
                'remaining_quantity': 'sum'
            }
            
            for col in warehouse_processed.columns:
                if col.startswith('Quantity '):
                    agg_dict[col] = 'sum'
            
            warehouse_processed = warehouse_processed.groupby('article_key').agg(agg_dict).reset_index()
        
        return warehouse_processed
    
    def _process_sales_data(self, sales_df: pd.DataFrame) -> pd.DataFrame:
        sales_df = self.clean_dataframe(sales_df)
        sales_renamed = self.rename_columns(sales_df, self.get_sales_column_mapping())
        
        sales_clean = sales_renamed.dropna(subset=['Article Description'])
        
        sales_processed = pd.DataFrame()
        sales_processed['company_code'] = sales_clean['Company Code']
        sales_processed['department'] = sales_clean['Department Name']
        sales_processed['category'] = sales_clean['Hierarchy Level 1 Description']
        sales_processed['nesto_mc'] = sales_clean['Nesto MC Description']
        sales_processed['brand'] = sales_clean['Brand Description']
        sales_processed['article_code'] = sales_clean['Article Code']
        sales_processed['article_description'] = sales_clean['Article Description']
        sales_processed['article_key'] = sales_clean['Article Code'].astype(str)
        
        stock_col = self.find_quantity_column(sales_clean, "sales_stock")
        sales_col = self.find_quantity_column(sales_clean, "sales_qty")
        
        if stock_col:
            sales_processed['current_stock'] = pd.to_numeric(sales_clean[stock_col], errors='coerce').fillna(0)
            print(f"Found sales stock column: {stock_col}")
        else:
            sales_processed['current_stock'] = 0
            print("Current stock column not found, defaulting to 0")
        
        if sales_col:
            sales_processed['sales_qty'] = pd.to_numeric(sales_clean[sales_col], errors='coerce').fillna(0)
            print(f"Found sales quantity column: {sales_col}")
        else:
            sales_processed['sales_qty'] = 0
            print("Sales quantity column not found, defaulting to 0")
        
        sales_processed = sales_processed[
            (sales_processed['article_code'].notna()) & 
            (sales_processed['article_code'] != '')
        ]
        
        sales_processed['total_grn'] = sales_processed['sales_qty'] + sales_processed['current_stock']
        sales_processed['sales_conversion'] = np.where(
            sales_processed['total_grn'] > 0,
            (sales_processed['sales_qty'] / sales_processed['total_grn']) * 100,
            0
        )
        
        return sales_processed
    
        # In stock_distribution.py, replace the _calculate_mc_performance method:

    def _calculate_mc_performance(self, sales_df: pd.DataFrame, start_date: str, end_date: str) -> Dict[str, Dict]:
        mc_performance = {}
        
        # Fix the date validation logic
        if (start_date is None or end_date is None or 
            not start_date or not end_date or 
            str(start_date).strip() in ['None', '', 'null'] or 
            str(end_date).strip() in ['None', '', 'null']):
            duration_days = 30
        else:
            try:
                start_date_str = str(start_date).strip()
                end_date_str = str(end_date).strip()
                start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
                duration_days = (end_dt - start_dt).days + 1
            except Exception as e:
                duration_days = 30
        
        for (company_code, nesto_mc), group in sales_df.groupby(['company_code', 'nesto_mc']):
            total_sales = group['sales_qty'].sum()
            current_stock = group['current_stock'].sum()
            daily_avg = total_sales / duration_days if duration_days > 0 else 0
            days_of_stock = (current_stock / daily_avg) if daily_avg > 0 else float('inf')
            
            total_grn = group['total_grn'].sum()
            mc_conversion = (total_sales / total_grn * 100) if total_grn > 0 else 0
            
            mc_performance[f"{company_code}_{nesto_mc}"] = {
                'company_code': company_code,
                'nesto_mc': nesto_mc,
                'daily_avg_sales': daily_avg,
                'current_stock': current_stock,
                'total_sales': total_sales,
                'days_of_stock': days_of_stock,
                'mc_conversion': mc_conversion,
                'total_grn': total_grn,
                'duration_days': duration_days
            }
        
        return mc_performance
    
    def _check_60_day_limit(self, current_stock: float, recommended_qty: int, daily_avg_sales: float) -> bool:
        if daily_avg_sales <= 0:
            return True
        
        total_stock_after = current_stock + recommended_qty
        days_of_stock = total_stock_after / daily_avg_sales
        
        return days_of_stock <= 60
    
    def _calculate_existing_distributions(self, warehouse_df: pd.DataFrame, sales_df: pd.DataFrame, 
                                        mc_performance: Dict[str, Dict]) -> List[Dict]:
        print("Calculating existing articles distribution...")
        
        high_conversion_items = sales_df[sales_df['sales_conversion'] >= 80].copy()
        print(f"High conversion items found: {len(high_conversion_items)}")
        
        warehouse_lookup = warehouse_df.set_index('article_key').to_dict('index')
        
        distribution_results = []
        
        for _, item in high_conversion_items.iterrows():
            article_key = item['article_key']
            
            if article_key not in warehouse_lookup:
                continue
                
            warehouse_stock = warehouse_lookup[article_key]
            mc_key = f"{item['company_code']}_{item['nesto_mc']}"
            mc_data = mc_performance.get(mc_key, {'mc_conversion': 0, 'total_sales': 0, 'daily_avg_sales': 0})
            
            if mc_data['mc_conversion'] > 60:
                recommended_qty = min(12, warehouse_stock['remaining_quantity'])
                
                if self._check_60_day_limit(item['current_stock'], recommended_qty, mc_data['daily_avg_sales']):
                    if recommended_qty > 0:
                        distribution_results.append({
                            'warehouse_code': warehouse_stock['company_code'],
                            'article_code': warehouse_stock['article_code'],
                            'article_description': warehouse_stock['article_description'],
                            'nesto_mc': item['nesto_mc'],
                            'brand': item['brand'],
                            'site_code': item['company_code'],
                            'article_sales_conversion': round(item['sales_conversion'], 2),
                            'article_sales_qty': int(item['sales_qty']),
                            'mc_conversion': round(mc_data['mc_conversion'], 2),
                            'warehouse_total_qty': warehouse_stock['total_quantity'],
                            'distribution_qty': recommended_qty,
                            'current_site_stock': item['current_stock'],
                            'distribution_type': 'Existing Article',
                            'daily_avg_sales': round(mc_data['daily_avg_sales'], 2),
                            'mc_total_sales': int(mc_data['total_sales']),
                            'warehouse_data': warehouse_stock
                        })
        
        final_distribution = self._resolve_conflicts_with_stock_tracking(distribution_results, warehouse_df)
        
        print(f"Final existing articles distribution: {len(final_distribution)}")
        return final_distribution
    
    def _resolve_conflicts_with_stock_tracking(self, distribution_results: List[Dict], warehouse_df: pd.DataFrame) -> List[Dict]:
        distribution_by_article = defaultdict(list)
        for dist in distribution_results:
            distribution_by_article[dist['article_code']].append(dist)
        
        final_distribution = []
        
        for article_code, sites in distribution_by_article.items():
            if len(sites) == 1:
                final_distribution.append(sites[0])
            else:
                sites.sort(key=lambda x: (x['article_sales_qty'], x['article_sales_conversion']), reverse=True)
                
                warehouse_qty = sites[0]['warehouse_total_qty']
                remaining_qty = warehouse_qty
                
                for site in sites:
                    if remaining_qty <= 0:
                        break
                    
                    allocated_qty = min(site['distribution_qty'], remaining_qty)
                    if allocated_qty > 0:
                        site['distribution_qty'] = allocated_qty
                        final_distribution.append(site)
                        remaining_qty -= allocated_qty
                        print(f"Allocated {allocated_qty} of {article_code} to {site['site_code']} (remaining: {remaining_qty})")
        
        return final_distribution
    
    def _update_warehouse_stock(self, warehouse_df: pd.DataFrame, distributions: List[Dict]) -> pd.DataFrame:
        warehouse_updated = warehouse_df.copy()
        
        distributed_by_article = defaultdict(int)
        for dist in distributions:
            distributed_by_article[dist['article_code']] += dist['distribution_qty']
        
        for article_code, total_distributed in distributed_by_article.items():
            mask = warehouse_updated['article_code'] == article_code
            if mask.any():
                current_remaining = warehouse_updated.loc[mask, 'remaining_quantity'].iloc[0]
                new_remaining = max(0, current_remaining - total_distributed)
                warehouse_updated.loc[mask, 'remaining_quantity'] = new_remaining
                print(f"Updated {article_code}: {current_remaining} -> {new_remaining} (distributed: {total_distributed})")
        
        return warehouse_updated
    
    def _calculate_mc_stock_after_existing(self, sales_df: pd.DataFrame, existing_distributions: List[Dict], 
                                         mc_performance: Dict[str, Dict]) -> Dict[str, Dict]:
        mc_stock_status = {}
        
        distributed_by_mc = defaultdict(int)
        for dist in existing_distributions:
            mc_key = f"{dist['site_code']}_{dist['nesto_mc']}"
            distributed_by_mc[mc_key] += dist['distribution_qty']
        
        for mc_key, perf in mc_performance.items():
            distributed_qty = distributed_by_mc.get(mc_key, 0)
            stock_after_distribution = perf['current_stock'] + distributed_qty
            days_after_distribution = (stock_after_distribution / perf['daily_avg_sales']) if perf['daily_avg_sales'] > 0 else float('inf')
            
            mc_stock_status[mc_key] = {
                'company_code': perf['company_code'],
                'nesto_mc': perf['nesto_mc'],
                'current_stock': perf['current_stock'],
                'distributed_qty': distributed_qty,
                'stock_after_distribution': stock_after_distribution,
                'days_after_distribution': days_after_distribution,
                'daily_avg_sales': perf['daily_avg_sales'],
                'needs_more_stock': days_after_distribution < 60 and perf['daily_avg_sales'] > 0
            }
        
        return mc_stock_status
    
    def _calculate_mc_stock_after_distribution(self, mc_stock_status: Dict[str, Dict], 
                                             new_distributions: List[Dict], mc_performance: Dict[str, Dict]) -> Dict[str, Dict]:
        additional_distributed = defaultdict(int)
        for dist in new_distributions:
            mc_key = f"{dist['site_code']}_{dist['nesto_mc']}"
            additional_distributed[mc_key] += dist['distribution_qty']
        
        updated_status = mc_stock_status.copy()
        for mc_key, additional_qty in additional_distributed.items():
            if mc_key in updated_status:
                updated_status[mc_key]['distributed_qty'] += additional_qty
                updated_status[mc_key]['stock_after_distribution'] += additional_qty
                
                daily_avg = updated_status[mc_key]['daily_avg_sales']
                if daily_avg > 0:
                    updated_status[mc_key]['days_after_distribution'] = updated_status[mc_key]['stock_after_distribution'] / daily_avg
                    updated_status[mc_key]['needs_more_stock'] = updated_status[mc_key]['days_after_distribution'] < 60
        
        return updated_status
    
    def _calculate_entirely_new_distributions(self, warehouse_df: pd.DataFrame, mc_stock_status: Dict[str, Dict],
                                            mc_performance: Dict[str, Dict], sales_df: pd.DataFrame) -> List[Dict]:
        print("Calculating entirely new articles distribution...")
        
        needy_mcs = {mc_key: status for mc_key, status in mc_stock_status.items() if status['needs_more_stock']}
        print(f"MCs needing more stock: {len(needy_mcs)}")
        
        if not needy_mcs:
            print("No MCs need additional stock, skipping entirely new articles")
            return []
        
        sold_articles = set(sales_df['article_key'].unique())
        
        entirely_new_articles = warehouse_df[~warehouse_df['article_key'].isin(sold_articles)]
        print(f"Entirely new articles found: {len(entirely_new_articles)}")
        
        distribution_results = []
        
        for idx, warehouse_item in entirely_new_articles.iterrows():
            if warehouse_item['remaining_quantity'] <= 0:
                continue
                
            for mc_key, mc_status in needy_mcs.items():
                if mc_status['nesto_mc'] == warehouse_item['nesto_mc']:
                    recommended_qty = min(12, warehouse_item['remaining_quantity'])
                    
                    if self._check_60_day_limit(mc_status['stock_after_distribution'], recommended_qty, mc_status['daily_avg_sales']):
                        if recommended_qty > 0:
                            distribution_results.append({
                                'warehouse_code': warehouse_item['company_code'],
                                'article_code': warehouse_item['article_code'],
                                'article_description': warehouse_item['article_description'],
                                'nesto_mc': warehouse_item['nesto_mc'],
                                'brand': warehouse_item['brand'],
                                'site_code': mc_status['company_code'],
                                'article_sales_conversion': 0,
                                'article_sales_qty': 0,
                                'mc_conversion': 0,
                                'warehouse_total_qty': warehouse_item['total_quantity'],
                                'distribution_qty': recommended_qty,
                                'current_site_stock': 0,
                                'distribution_type': 'Entirely New Article',
                                'daily_avg_sales': round(mc_status['daily_avg_sales'], 2),
                                'mc_total_sales': int(mc_performance.get(mc_key, {}).get('total_sales', 0)),
                                'warehouse_data': warehouse_item.to_dict()
                            })
                            
                            warehouse_df.loc[idx, 'remaining_quantity'] -= recommended_qty
                            if warehouse_df.loc[idx, 'remaining_quantity'] <= 0:
                                break
        
        print(f"Entirely new articles distribution: {len(distribution_results)}")
        return distribution_results
    
    def _calculate_site_new_distributions(self, warehouse_df: pd.DataFrame, sales_df: pd.DataFrame,
                                        mc_stock_status: Dict[str, Dict], mc_performance: Dict[str, Dict]) -> List[Dict]:
        print("Calculating site-wise new articles distribution...")
        
        needy_mcs = {mc_key: status for mc_key, status in mc_stock_status.items() if status['needs_more_stock']}
        print(f"MCs still needing more stock: {len(needy_mcs)}")
        
        if not needy_mcs:
            print("No MCs need additional stock, skipping site-wise new articles")
            return []
        
        site_articles = {}
        for _, row in sales_df.iterrows():
            site_key = f"{row['company_code']}_{row['nesto_mc']}"
            if site_key not in site_articles:
                site_articles[site_key] = set()
            site_articles[site_key].add(row['article_key'])
        
        distribution_results = []
        
        for idx, warehouse_item in warehouse_df.iterrows():
            if warehouse_item['remaining_quantity'] <= 0:
                continue
                
            article_key = warehouse_item['article_key']
            
            for mc_key, mc_status in needy_mcs.items():
                if (mc_status['nesto_mc'] == warehouse_item['nesto_mc'] and 
                    mc_key in site_articles and
                    article_key not in site_articles[mc_key]):
                    
                    recommended_qty = min(12, warehouse_item['remaining_quantity'])
                    
                    if self._check_60_day_limit(mc_status['stock_after_distribution'], recommended_qty, mc_status['daily_avg_sales']):
                        if recommended_qty > 0:
                            distribution_results.append({
                                'warehouse_code': warehouse_item['company_code'],
                                'article_code': warehouse_item['article_code'],
                                'article_description': warehouse_item['article_description'],
                                'nesto_mc': warehouse_item['nesto_mc'],
                                'brand': warehouse_item['brand'],
                                'site_code': mc_status['company_code'],
                                'article_sales_conversion': 0,
                                'article_sales_qty': 0,
                                'mc_conversion': 0,
                                'warehouse_total_qty': warehouse_item['total_quantity'],
                                'distribution_qty': recommended_qty,
                                'current_site_stock': 0,
                                'distribution_type': 'Site New Article',
                                'daily_avg_sales': round(mc_status['daily_avg_sales'], 2),
                                'mc_total_sales': int(mc_performance.get(mc_key, {}).get('total_sales', 0)),
                                'warehouse_data': warehouse_item.to_dict()
                            })
                            
                            warehouse_df.loc[idx, 'remaining_quantity'] -= recommended_qty
                            if warehouse_df.loc[idx, 'remaining_quantity'] <= 0:
                                break
        
        print(f"Site-wise new articles distribution: {len(distribution_results)}")
        return distribution_results
    
    def _create_output_data(self, all_distributions: List[Dict], existing_distributions: List[Dict],
                          entirely_new_distributions: List[Dict], site_new_distributions: List[Dict],
                          warehouse_df: pd.DataFrame, sales_df: pd.DataFrame,
                          start_date: str, end_date: str, mc_performance: Dict[str, Dict]) -> Dict[str, pd.DataFrame]:
        
        pivot_data = self._create_pivot_table(all_distributions)
        detailed_data = self._create_detailed_table(all_distributions)
        summary_data = self._create_summary_analysis(
            existing_distributions, entirely_new_distributions, site_new_distributions,
            warehouse_df, sales_df, start_date, end_date, mc_performance
        )
        
        return {
            'Distribution Pivot': pivot_data,
            'Detailed Distribution': detailed_data,
            'Summary Analysis': summary_data
        }
    
    def _create_pivot_table(self, all_distributions: List[Dict]) -> pd.DataFrame:
        pivot_data = []
        distribution_grouped = defaultdict(list)
        
        for item in all_distributions:
            key = f"{item['warehouse_code']}_{item['article_code']}"
            distribution_grouped[key].append(item)
        
        for key, distributions in distribution_grouped.items():
            warehouse_code, article_code = key.split('_', 1)
            
            pivot_row = {
                'Warehouse Code': warehouse_code,
                'Warehouse Total Quantity': distributions[0]['warehouse_total_qty'],
                'Article Code': distributions[0]['article_code'],
                'Article Description': distributions[0]['article_description'],
                'Nesto MC': distributions[0]['nesto_mc'],
                'Brand': distributions[0]['brand'],
                'Distribution Type': distributions[0]['distribution_type']
            }
            
            if 'warehouse_data' in distributions[0]:
                warehouse_data = distributions[0]['warehouse_data']
                for col in warehouse_data:
                    if col.startswith('Quantity '):
                        pivot_row[col] = warehouse_data[col]
            
            for dist in distributions:
                pivot_row[dist['site_code']] = dist['distribution_qty']
            
            pivot_data.append(pivot_row)
        
        return pd.DataFrame(pivot_data)
    
    def _create_detailed_table(self, all_distributions: List[Dict]) -> pd.DataFrame:
        detailed_data = []
        
        for item in all_distributions:
            detail_row = {
                'Warehouse Code': item['warehouse_code'],
                'Article Code': item['article_code'],
                'Article Description': item['article_description'],
                'Nesto MC': item['nesto_mc'],
                'Brand': item['brand'],
                'Site Code': item['site_code'],
                'Distribution Type': item['distribution_type'],
                'Warehouse Total Qty': item['warehouse_total_qty'],
                'Distribution Qty': item['distribution_qty'],
                'Article Sales Conversion %': item['article_sales_conversion'],
                'Article Sales Qty': item['article_sales_qty'],
                'MC Conversion %': item.get('mc_conversion', 0),
                'Current Site Stock': item['current_site_stock'],
                'Daily Avg Sales': item.get('daily_avg_sales', 0),
                'MC Total Sales': item.get('mc_total_sales', 0)
            }
            
            if 'warehouse_data' in item:
                warehouse_data = item['warehouse_data']
                for col in warehouse_data:
                    if col.startswith('Quantity '):
                        detail_row[col] = warehouse_data[col]
            
            detailed_data.append(detail_row)
        
        return pd.DataFrame(detailed_data)
    
    def _create_summary_analysis(self, existing_distributions: List[Dict], entirely_new_distributions: List[Dict],
                               site_new_distributions: List[Dict], warehouse_df: pd.DataFrame,
                               sales_df: pd.DataFrame, start_date: str, end_date: str, 
                               mc_performance: Dict[str, Dict]) -> pd.DataFrame:
        
        try:
            if start_date and end_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                duration_days = (end_dt - start_dt).days + 1
            else:
                duration_days = 30
        except:
            duration_days = 30
        
        summary_data = []
        
        distributions_by_mc = defaultdict(lambda: {'existing': 0, 'entirely_new': 0, 'site_new': 0, 'total_qty': 0})
        
        for dist in existing_distributions:
            mc_key = f"{dist['site_code']}_{dist['nesto_mc']}"
            distributions_by_mc[mc_key]['existing'] += 1
            distributions_by_mc[mc_key]['total_qty'] += dist['distribution_qty']
        
        for dist in entirely_new_distributions:
            mc_key = f"{dist['site_code']}_{dist['nesto_mc']}"
            distributions_by_mc[mc_key]['entirely_new'] += 1
            distributions_by_mc[mc_key]['total_qty'] += dist['distribution_qty']
        
        for dist in site_new_distributions:
            mc_key = f"{dist['site_code']}_{dist['nesto_mc']}"
            distributions_by_mc[mc_key]['site_new'] += 1
            distributions_by_mc[mc_key]['total_qty'] += dist['distribution_qty']
        
        for mc_key, perf in mc_performance.items():
            dist_data = distributions_by_mc[mc_key]
            
            stock_after_distribution = perf['current_stock'] + dist_data['total_qty']
            days_after_distribution = stock_after_distribution / perf['daily_avg_sales'] if perf['daily_avg_sales'] > 0 else float('inf')
            
            actual_duration = perf.get('duration_days', duration_days)
            
            summary_row = {
                'Site Code': perf['company_code'],
                'Nesto MC': perf['nesto_mc'],
                'Analysis Period (Days)': actual_duration,
                'Total Sales in Period': int(perf['total_sales']),
                'Daily Average Sales': round(perf['daily_avg_sales'], 2),
                'Current Stock': int(perf['current_stock']),
                'Current Days of Stock': round(perf['days_of_stock'], 1) if perf['days_of_stock'] != float('inf') else 'No Sales',
                'Existing Articles Distributed': dist_data['existing'],
                'Entirely New Articles Distributed': dist_data['entirely_new'],
                'Site New Articles Distributed': dist_data['site_new'],
                'Total Recommended Distribution': dist_data['total_qty'],
                'Stock After Distribution': int(stock_after_distribution),
                'Days After Distribution': round(days_after_distribution, 1) if days_after_distribution != float('inf') else 'No Sales',
                'Within 60 Days Limit': 'Yes' if days_after_distribution <= 60 else 'No' if days_after_distribution != float('inf') else 'No Sales',
                'Distribution Status': 'Active' if dist_data['total_qty'] > 0 else 'No Distribution'
            }
            
            summary_data.append(summary_row)
        
        summary_df = pd.DataFrame(summary_data)
        if not summary_df.empty:
            summary_df = summary_df.sort_values('Total Recommended Distribution', ascending=False)
        
        return summary_df
    
    def _save_excel_with_colors(self, output_data: Dict[str, pd.DataFrame], output_path: str) -> None:
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in output_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                if 'Distribution Pivot' in output_data:
                    self._apply_color_coding(writer.book['Distribution Pivot'], output_data['Distribution Pivot'])
                
                if 'Detailed Distribution' in output_data:
                    self._apply_color_coding(writer.book['Detailed Distribution'], output_data['Detailed Distribution'])
                
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
        
        except Exception as e:
            raise FileProcessingError(f"Error saving Excel file to {output_path}: {str(e)}")
    
    def _apply_color_coding(self, worksheet, df: pd.DataFrame) -> None:
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        quantity_columns = {}
        for col_idx, col_name in enumerate(df.columns, 1):
            if str(col_name).startswith('Quantity '):
                quantity_columns[col_idx] = col_name
        
        for row_idx in range(2, len(df) + 2):
            for col_idx, col_name in quantity_columns.items():
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if cell.value is not None and cell.value != '' and cell.value != 0:
                    if 'Quantity 0-30' in col_name or 'Quantity 31-60' in col_name:
                        cell.fill = green_fill
                    elif 'Quantity 61-90' in col_name:
                        cell.fill = orange_fill
                    elif 'Quantity 91-180' in col_name:
                        cell.fill = yellow_fill
                    elif ('Quantity 181-365' in col_name or 'Quantity >365' in col_name or 'Quantity >180' in col_name):
                        cell.fill = red_fill